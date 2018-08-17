import openpyxl

class A:
    def read(self):
        e1=openpyxl.load_workbook("C:/Users/meridian/Desktop/提取练习/提取.xlsx")

        w1=e1["Sheet1"]
        print(w1['A1'].value)

        #最大行数和列数
        Maxr=w1.max_row
        Maxc=w1.max_column
        print(Maxc,Maxr)

        # 此列表用来存所有行中的数据
        l1=list()
        for i in w1['B']:
            #print(i.value,end="\n")
            l1.append(i.value)

        #print(len(l1))
        # 此列表用来存拆分后的数据
        l2=list()
        for j in l1:
            if j[0]=="。":
                j=j[1:]
            if j[0]=="肝":
                l3=j.split('。')
                for i in l3:
                    if i=="":
                        l3.remove(i)
                        # if i[0].isalpha():
                        #     l3.remove(i)
                    l2.append(l3)
        d = 1
        l4 = list()
        t1=""

        for i in l2:

            #l22.append()
            for m in i:
                if m!="":
                    if len(m) >2:
                        if m[0] =='甲':
                            t= str(w1['A{0}'.format(d)].value) + m

                            l4.append(t)
            d=d+1


        # 开始写入

        tq=openpyxl.Workbook()
        Sheet1=tq.active
        Sheet1.title="已提取"
        #l5=list()
        print("开始写入")

        # 行数很少时可以使用
        #Sheet1.append(l4)

        l=len(l4)
        m=1
        for i in l4:

            Sheet1['A{0}'.format(m)]=i
            m=m+1

        print("写入完成")

        tq.save('C:/Users/meridian/Desktop/提取练习/已提取.xlsx')



if __name__ == '__main__':
    a1=A()
    a1.read()