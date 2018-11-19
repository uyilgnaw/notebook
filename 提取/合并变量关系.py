import openpyxl
import xlwt
def read():
    e1 = openpyxl.load_workbook("C:/Users/meridian/Desktop/提取练习/待合并.xlsx")
    w1 = e1.active

    Maxr = w1.max_row
    Maxc = w1.max_column

    print('最大行数',Maxr)
    print('最大列数',Maxc)
    # 包含去重后的标准名称
    s1 = set()
    for i in w1['B']:

        s1.add(i.value)
        # print(i.value)

    # print(len(s1))
    l2 =list()
    for i in w1.rows:
        l1 = list()
        for j in i:
            # print(j.value)
            l1.append(j.value)
        l2.append(l1)

    print(l2)

    str = '^'
    l3 = list()
    for i in l2:
        # print(i)
        # print(str.join(i))
        l3.append(str.join(i))
    # print(l3)

    d1 = dict()
    for i in s1:
        l4 = list()
        for j in l3:
            # print(j[j.find('^') + 1:])
            if i == j[j.find('^')+1:]:
                if j[:j.find('^')] in l4:
                    continue
                else:
                    l4.append(j[:j.find('^')])
        d1[i] = l4





    file = openpyxl.Workbook()
    table = file.active
    table.cell(row=1,column=1).value = "标准名"
    table.cell(row=1,column=2).value = "原始指标名"
    for i in d1.keys():
        d1[i].append(i)
    for i in d1.values():
        i.reverse()
        table.append(i)
    file.save("C:/Users/meridian/Desktop/提取练习/已合并.xlsx")


if __name__ == '__main__':
    read()