import openpyxl


class a:

    def read(self):
        t1=openpyxl.load_workbook(filename="C:/Users/meridian/Desktop/提取练习/提取.xlsx")
        # 打印该工作簿中的活动工作表
        print(t1.active)
        e1=t1['Sheet1']
        print(e1['A1'].value)
        Maxc=e1.max_column
        Maxr=e1.max_row
        print('最大列数为', Maxc)
        print('最大行数为', Maxr)
        l1=list()
        for i in e1['A']:
            #print(i.value,end="\n")
            # 将非空单元格添加到列表当中
            if i.value is not  None:
                l1.append(i.value)
        # 查看非空单元格有多少
        #print(len(l1))
        #for j in l1:
        #print(l1[1])
        c0=l1[1]
        c1=c0.replace("_x000D_","")
        #print(type(c1))
        # c2=c1.find("_x000D_")
        # c11=c1[0:c2]
        # print(c11)
        #c3=c1.find("_x000D_",55)
        #print(c3)
        s1=0
        l2=list()
        while c1.find("\n",s1):
            s2=c1.find("\n",s1)
            c2=c1[s1:s2]
            l2.append(c2)
            if s2<0:
                break
            else:
                s1=s2+1
        print(len(l2))

        #print(c1)





if __name__ == '__main__':
    a1=a()
    a1.read()





