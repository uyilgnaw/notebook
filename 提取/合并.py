import openpyxl

def read():
    e1=openpyxl.load_workbook("C:/Users/meridian/Desktop/提取练习/待合并.xlsx")
    w1=e1.active
    Maxr=w1.max_row
    Maxc=w1.max_column
    print("最大列数为：",Maxc,"最大行数为：",Maxr)
    #
    # for i in w1['C']:
    #     if i.value is not None and i.value[0] != "#":
    #         print(i.value)

    l2=list()
    for i in w1.rows:
        i=list(i)
        #print(i)
        l1 = list()
        for j in i:

            if j.value is not None and j.value[0] != "#":
                #print(j.value)
                l1.append(j.value)
        #print(len(l1))
        # if len(l1) == 0:
        #     l1.append("暂未对应")

        l2.append(l1)
        # if l1 != "":
        #     l2.append(l1)

    print(len(l2))
    l4=list()
    for i in l2:
        # print(i)
        # print(len(i))
        l3 = list()
        for j in range(len(i)):
            if i[j][0]=='T':
                l3.append(i[j] + "-" + i[j+1])
                #print(i[j] + "-" + i[j+1] )
        l4.append(l3)
    # 开始合并
    print(len(l4))

    # 使用join函数将list列表中的数据进行连接
    l6=list()
    str="；"
    for i in l4:
        l5 = list()

        l5.append(str.join(i))


        l6.append(l5)

    print(len(l6))


    # 开始写入
    e2=openpyxl.Workbook()
    w2=e2.active

    p=0
    for i in l6:
        p=p+1
        w2.append(i)
        print('已经合并了{0}行'.format(p))

    print('合并完成')
    print('正在保存')
    e2.save("C:/Users/meridian/Desktop/提取练习/已合并.xlsx")
    print('保存完成')


read()