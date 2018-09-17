import openpyxl
import datetime
# 此程序可处理超过20万行的工作表。
class A():
    def read(self):
        e1=openpyxl.load_workbook(filename='C:/Users/meridian/Desktop/提取练习/待提取.xlsx',read_only=True)
        w1=e1.active

        Maxr=w1.max_row
        Maxc=w1.max_column
        print('最大行数为：',Maxr)
        print('最大列数为:',Maxc)
        l1=list(w1.rows)
        l2=list()
        l3=list()

        l5=list()
        # 下面这个for循环是将重复数据筛选出来
        for i in w1.rows:
            #print(i[0].value)
            if i[0].value not in l2:
                l2.append(i[0].value)
            else:
                for m in l1:
                    if i[0].value==m[0].value:
                        l3.append(m)
        # 这个for循环是将筛选出来的重复数据去重
        l6=list()
        for i in l3:
            if i not in l6:
                l6.append(i)


        # 这个for循环是将每一行重复数据形成一个列表。添加到最终的l5列表当中
        for i in l6:
            l4=list()
            for j in i:
                l4.append(j.value)
            l5.append(l4)

        #开始写入
        e2=openpyxl.Workbook()
        w2=e2.active
        p=0
        # 这个for循环是将l5中的数据按行写入到新的工作簿中
        for i in l5:
            w2.append(i)
            p=p+1

            print('已经找到写入了{0}'.format(p))

        e2.save("C:/Users/meridian/Desktop/提取练习/已提取.xlsx")





if __name__ == '__main__':
    # 记录程序开始时间
    starttime=datetime.datetime.now()

    a=A()
    a.read()
    # 记录程序结束时间
    endtime=datetime.datetime.now()
    # 记录程序运行了多少秒
    print(endtime-starttime,'秒')