# -*- encoding: utf-8 -*-
"""
excel2excel5.py
Created on 2018/1/2 17:17
Copyright (c) 2018/1/2, 
@author: 小马同学(majstx@163.com)
"""

# 已对结果进行对应的格式化(清除不可见字符、去掉换行符)
# 处理隐藏的信息

import myutil
import openpyxl
import os
import time
import re


def jg2xlsx():
    # 人的信息
    peinfo = []
    tjbh = file.split('_')[0]
    khbh = file.split('-')[0]
    xm = file.split('_')[1].split('.')[0]

    xbnl = sheetd.HeaderFooter.oddHeader.right.text
    pos_xb = xbnl.find('性别：')
    xbnl_list = xbnl[pos_xb:].split()
    xb = xbnl_list[0].split('：')[1]
    nl = xbnl_list[1].split('：')[1]
    nl = myutil.confir(nl)
    try:
        nl = re.findall(r"\d+\.?\d*", nl)[0]
    except:
        nl = ''
    peinfo.append(tjbh);peinfo.append(khbh);peinfo.append(xm)
    peinfo.append(xb);peinfo.append(nl)

    for col_num in range(1, len(peinfo)+1):
        cell_num = myutil.number2char(col_num) + str(row_num)
        col_num -= 1
        value = peinfo[col_num]
        sheet[cell_num] = value  # 添加个人信息
        sheet1[cell_num] = value
        sheet2[cell_num] = value

    # 项目的信息
    for i in range(1, nrows+1):
        flag = sheetd.cell(row=i, column=1).value
        if  flag == '项目名称':
            ksvalue = sheetd.cell(row=i-2, column=1).value
            if ksvalue == '小  结：': #
                ksvalue = lastksvalue
            else:
                ksvalue = sheetd.cell(row=i - 2, column=1).value
            dxvalue = sheetd.cell(row=i-1, column=1).value
            # 小项及体检结果
            for r in range(i+1, nrows+1):   # 行
                xxvalue = sheetd.cell(row=r, column=1).value
                if xxvalue == '小  结：':
                    break
                index = ksvalue + '_' + dxvalue + '_' + xxvalue
                jgvalue = sheetd.cell(row=r, column=2).value
                jgvalue = myutil.str_formatxm(jgvalue)
                dwvalue = sheetd.cell(row=r, column=3).value
                fwvalue = sheetd.cell(row=r, column=4).value
                fwdwvalue = fwvalue + '[' + dwvalue + ']'
                if fwdwvalue == '[]':
                    fwdwvalue = ''
                elif fwdwvalue == '-[]':
                    fwdwvalue = ''

                # 数值分类  判断该结果是正常？偏高？偏低？
                if myutil.is_number(jgvalue):
                    try:
                        fl_value = myutil.compare(jgvalue,fwvalue)
                    except:
                        fl_value = jgvalue
                else:
                    fl_value = jgvalue

                # 如果指标在字典中就取出该列，不在添加表头
                if index in dict_allindexjg:
                    col_num = dict_allindexjg[index]
                else:
                    col_num = len(dict_allindexjg) + 1
                    dict_allindexjg[index] = col_num  # 添加进字典
                    cell_bt = myutil.number2char(col_num) + '1'
                    sheet[cell_bt] = index
                    sheet1[cell_bt] = index
                    sheet2[cell_bt] = index
                # 定位此时单元格的位置
                cell_num = myutil.number2char(col_num) + str(row_num)

                # 对结果按照预定的规则格式化
                sheet[cell_num] = jgvalue
                sheet1[cell_num] = fl_value
                sheet2[cell_num] = fwdwvalue

            lastksvalue = ksvalue


def xj2xlsx():

    # 人的信息
    peinfo = []
    tjbh = file.split('_')[0]
    peinfo.append(tjbh)

    for col_num in range (1,len(peinfo)+1):
        cell_num = myutil.number2char(col_num) + str(row_num)
        col_num -= 1
        value = peinfo[col_num]
        sheet3[cell_num] = value # 添加个人信息
        # sheet2[cell_num] = value

    # 项目的信息
    for i in range(1, nrows+1):
        flag = sheetd.cell(row=i, column=1).value
        if  flag == '项目名称':
            ksvalue = sheetd.cell(row=i-2, column=1).value
            if ksvalue == '小  结：':
                ksvalue = lastksvalue
            else:
                ksvalue = sheetd.cell(row=i-2, column=1).value
            dxvalue = sheetd.cell(row=i-1, column=1).value
            # 小项及体检结果
            for r in range(i+1, nrows+1):   # 行
                xxvalue = sheetd.cell(row=r, column=1).value
                if xxvalue == '小  结：':
                    xxvalue = '小结'
                    index = ksvalue + '_' + dxvalue + '_' + xxvalue
                    xjvalue = sheetd.cell(row=r, column=2).value
                    xjvalue = myutil.str_formatxm(xjvalue)

                # 如果指标在字典中就取出该列，不在添加表头
                    if index in dict_allindexxj:
                        col_num = dict_allindexxj[index]
                    else:
                        col_num = len(dict_allindexxj) + 1
                        dict_allindexxj[index] = col_num  # 添加进字典
                        cell_bt = myutil.number2char(col_num) + '1'
                        sheet3[cell_bt] = index
                    # 定位此时单元格的位置
                    cell_num = myutil.number2char(col_num) + str(row_num)
                    try:
                        sheet3[cell_num] = xjvalue
                    except:
                        print('=======此处异常暂不处理=========')
                    break

            lastksvalue = ksvalue


def zs2xlsx():
    # 人的信息
    tjbh = file.split('_')[0]
    # 项目的信息
    for i in range(1, nrows+1):
        flag = sheetd.cell(row=i, column=1).value
        if flag == '综  述：':
            zsi = i
        if flag == '建  议：':
            jyi = i

    if not 'jyi' in locals().keys():   # 判断是否有建议，也就是看jyi是否定义
        # 综述
        zsvaluez = ''
        for i in range(zsi, jyi):
            zsvalue = sheetd.cell(row=i, column=2).value
            zsvaluez = zsvaluez + zsvalue
        # 建议
        jyvaluez = ''
    else:
        # 综述
        zsvaluez = ''
        for i in range(zsi, jyi):
            zsvalue = sheetd.cell(row=i, column=2).value
            # print type(zsvalue)
            if zsvalue is None:
                zsvalue = ''
            zsvaluez = zsvaluez + zsvalue
        # 建议
        jyvaluez = ''
        for i in range(jyi, nrows+1):
            jyvalue = sheetd.cell(row=i, column=2).value
            if jyvalue is None:
                jyvalue = ''
            jyvaluez = jyvaluez + jyvalue

    sheet4.cell(row=row_num, column=1).value = tjbh
    sheet4.cell(row=row_num, column=2).value = zsvaluez
    sheet4.cell(row=row_num, column=3).value = jyvaluez


if __name__ == '__main__':
    workbook = openpyxl.Workbook()
    workbook1 = openpyxl.Workbook()
    workbook2 = openpyxl.Workbook()
    workbook3 = openpyxl.Workbook()
    workbook4 = openpyxl.Workbook()
    sheet = workbook.active    # 原始表
    sheet1 = workbook1.active   # 分类表
    sheet2 = workbook2.active  # 范围表
    sheet3 = workbook3.active  # 小结
    sheet4 = workbook4.active  # 综述建议

    # 结果表
    pebt_listjg = ['体检编号', '客户编号', '姓名', '性别', '年龄']
    sheet.append(pebt_listjg)
    sheet1.append(pebt_listjg)
    sheet2.append(pebt_listjg)
    dict_allindexjg = dict(zip(pebt_listjg, range(1, len(pebt_listjg) + 1)))

    # 小结表
    pebt_listxj = ['体检编号']
    sheet3.append(pebt_listxj)
    dict_allindexxj = dict(zip(pebt_listxj, range(1, len(pebt_listxj) + 1)))

    # 综述建议
    pebt_listzj = ['体检编号', '综述', '建议']
    sheet4.append(pebt_listzj)

    path = "C:\\Users\\meridian\\Desktop\\襄阳总\\总\\"#源路径
    files = os.listdir(path)
    all_num = len(files)
    row_num = 2
    for file in files:
        if row_num % 5 == 0:
            print(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), u"【进度信息：%d >>>> %d 】" % (row_num, all_num))
        try:
            wbd = openpyxl.load_workbook(u"C:\\Users\\meridian\\Desktop\\襄阳总\\总\\%s"%file)
        except:
            print (time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), u"%s 该文件格式无效"%file)
            continue
        sheetd = wbd.get_active_sheet()
        # sheet_names = wbd.get_sheet_names()
        # sheetd = wbd.get_sheet_by_name(sheet_names[0])
        nrows = sheetd.max_row
        ncols = sheetd.max_column
        jg2xlsx()
        xj2xlsx()
        zs2xlsx()
        row_num += 1

    print('save source...')
    workbook.save("C:\\Users\\meridian\\Desktop\\襄阳out（最新）\\襄阳_原始.xlsx")
    print('save fl...')
    workbook1.save("C:\\Users\\meridian\\Desktop\\襄阳out（最新）\\襄阳_分类.xlsx")
    print('save fwdw...')
    workbook2.save("C:\\Users\\meridian\\Desktop\\襄阳out（最新）\\襄阳_范围.xlsx")
    print('save xj...')
    workbook3.save("C:\\Users\\meridian\\Desktop\\襄阳out（最新）\\襄阳_小结.xlsx")
    print('save zj...')
    workbook4.save("C:\\Users\\meridian\\Desktop\\襄阳out（最新）\\襄阳_综述建议.xlsx")
    print('Save completion')