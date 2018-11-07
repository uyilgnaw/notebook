# -*- encoding: utf-8 -*-
"""
xmldemo1.py
Created on 2018/8/22 17:15
@author: 马家树(majstx@163.com)
"""
import os
import xml.etree.ElementTree as ET
import openpyxl
import time

import myutil


def xml2xlsx():
    reps = tree.findall('Reports')
    global row_num
    # print(reps)
    for rs in reps:
        for rep in rs:  # 遍历出所有的体检单
    
            baseinfo = rep.find('BaseInfo')
            admid = baseinfo.find('AdmID').text
            patientid = baseinfo.find('PatientID').text
            patientname = baseinfo.find('PatientName').text
            sex = baseinfo.find('Sex').text
            age = baseinfo.find('Age').text
            try:
                workplace = baseinfo.find('GroupName').text
            except:
                workplace = ''
            checkdate = baseinfo.find('CheckDate').text
            try:
                tel = baseinfo.find('Tel').text
            except:
                tel = ''
            try:
                patid = baseinfo.find('PatID').text
            except:
                patid = ''
            # adD = baseinfo.find('AdmID').text
            adviceinfo = rep.find('Advice').find('AdviceInfo').text

            peinfo_list = [admid, patientid, patientname, sex, age, workplace, checkdate, tel, patid, adviceinfo]

            for col_num in range(1, len(peinfo_list) + 1):
                cell_num = myutil.number2char(col_num) + str(row_num)
                col_num -= 1
                value = peinfo_list[col_num]
                sheet[cell_num] = value  # 添加个人信息
                sheet1[cell_num] = value
                sheet2[cell_num] = value

            # 获取体检项目内容
            try:
                item = rep.find('TestItems').findall('TestItem')
            except:
                continue
            for i in item:  # TestItem
                dxmc = i.find('ItemName').text
                try:
                    itemdetail = i.find('ItemDetails').findall('ItemDetail')
                except:
                    continue
                for val in itemdetail:
                    xxmc = val.find('DetailName').text
                    try:
                        range_val = val.find('Range').text
                    except:
                        range_val = ''
                    try:
                        unit = val.find('Unit').text
                    except:
                        unit = ''
                    result = val.find('Result').text
                    result = myutil.str_formatxm(result)
                    indexval = dxmc + '_' + xxmc
                    fwdwval = range_val + '[' + unit + ']'
                    if unit == '':
                        fwdwval = range_val
                    if fwdwval == '[]':
                        fwdwval = ''

                    # 数值分类  判断该结果是正常？偏高？偏低？
                    if myutil.is_number(result):
                        try:
                            classifval = myutil.compare(result, range_val)
                        except:
                            classifval = result
                    else:
                        classifval = result

                    # 如果指标在字典中就取出该列，不在添加表头
                    if indexval in dict_allindex:
                        col_num = dict_allindex[indexval]
                    else:
                        col_num = len(dict_allindex) + 1
                        dict_allindex[indexval] = col_num  # 添加进字典
                        cell_bt = myutil.number2char(col_num) + '1'
                        sheet[cell_bt] = indexval
                        sheet1[cell_bt] = indexval
                        sheet2[cell_bt] = indexval

                    # 定位此时单元格的位置
                    cell_num = myutil.number2char(col_num) + str(row_num)
                    sheet[cell_num] = result
                    sheet1[cell_num] = classifval
                    sheet2[cell_num] = fwdwval
            row_num += 1


if __name__ == '__main__':

    workbook = openpyxl.Workbook()
    workbook1 = openpyxl.Workbook()
    workbook2 = openpyxl.Workbook()
    sheet = workbook.active  # 原始表
    sheet1 = workbook1.active  # 分类表
    sheet2 = workbook2.active  # 范围表

    pebt_list = ['编号', '体检编号', '姓名', '性别', '年龄', '工作单位', '体检日期', '联系电话', '身份证号', '建议']
    sheet.append(pebt_list)
    sheet1.append(pebt_list)
    sheet2.append(pebt_list)
    row_num = 2
    dict_allindex = dict(zip(pebt_list, range(1, len(pebt_list) + 1)))

    files = os.listdir('C:/Users/meridian/Desktop/江西-南昌市中西医结合医院-东华-xml/中国健康促进基金会-体检报告数据导出')
    file_len = len(files)
    num = 1
    for file in files:
        print(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), '一共是%s个当前处理的是第%s个' % (file_len, num))
        num += 1
        path = 'C:/Users/meridian/Desktop/江西-南昌市中西医结合医院-东华-xml/中国健康促进基金会-体检报告数据导出/%s' % file
        # print(path)
        try:
            tree = ET.parse(path)
        except:
            continue
        xml2xlsx()

    print(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), '开始保存...')
    workbook.save('C:/Users/meridian/Desktop/out/南昌市中西医结合医院_原始.xlsx')
    workbook1.save('C:/Users/meridian/Desktop/out/南昌市中西医结合医院_分类.xlsx')
    workbook2.save('C:/Users/meridian/Desktop/out/南昌市中西医结合医院_范围.xlsx')
    print(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), '保存成功！')